from src.excel_script_meta import ExcelScriptMeta
from openpyxl import Workbook


class ExcelScript(metaclass=ExcelScriptMeta):
    def __init__(self, output_fp):
        self.workbook = Workbook()
        self.output = output_fp

        # remove default worksheet created
        def_ws = self.workbook.get_sheet_by_name("Sheet")
        self.workbook.remove(def_ws)

    def create_worksheet(self, title):
        worksheet = self.workbook.create_sheet(title)
        return worksheet

    def setup_worksheet(self, worksheet, columns_list):
        # for now, creates necessary columns for each commenter.
        for i, col in enumerate(columns_list):
            worksheet.cell(row=1, column=i+1, value=col)

    def add_worksheet_contents(self, worksheet, contents, position):
        row = position[0]
        col = position[1]
        worksheet.cell(row=row, column=col, value=contents)
    
    def save_workbook(self):
        self.workbook.save(self.output)

    def delete_worksheet(self, worksheet):
        self.workbook.remove(worksheet)

    def get_nrows_in_col(self, worksheet, column):
        i = 0
        while True:
            i += 1
            cell = worksheet.cell(row=i, column=column)
            content = cell.value
            if content is None:
                nrows = i - 1
                break
        return nrows
