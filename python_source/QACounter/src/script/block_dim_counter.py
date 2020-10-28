import os
import math
import ezdxf
import ezdxf.math
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from ezdxf.tools.rgb import DXF_DEFAULT_COLORS, DXF_DEFAULT_COLORS_2, int2rgb
from src.util.constants import Constants
from src.cad.cad_manager import CadManager


class BlockDimCounter:
    def __init__(self):
        pass

    def iter_input(self):
        """
            gets every file in the input folder and process the necessary data
        """
        input_dir = Constants.INPUT_DIR
        for dir_path, dir_names, file_names in os.walk(input_dir):
            for file_name in file_names:
                file_full_path = os.path.join(dir_path, file_name)
                if file_full_path.endswith(Constants.DXF_FILE_EXT):
                    # for every dxf file you create 1 excel file
                    workbook = Workbook()
                    cad_manager = CadManager(file_full_path)
                    modelspace = cad_manager.get_modelspace()
                    # getting all the blocks
                    for block_ent in cad_manager.get_blockreference(modelspace):
                        # polyline container
                        pl_list = []
                        notch_list = []
                        # getting all the polylines in the block reference
                        for polyline in cad_manager.get_polylines(block_ent.virtual_entities()):
                            if polyline.closed:
                                found = False
                                count = 1
                                width, height = self.__compute_width_height(polyline, cad_manager.unit)
                                rgb_str = self.__get_rgb(polyline)
                                if not self.__is_notch(polyline):
                                    # checking if same polyline already exist in the list.
                                    for ent_member in pl_list:
                                        if ent_member['A'] == rgb_str and ent_member['B'] == width and \
                                                ent_member['C'] == height:
                                            ent_member['D'] += 1
                                            found = True
                                            break
                                    if not found:
                                        pl_data_dict = self.__compose_data_dict(rgb_str, width, height, count)
                                        pl_list.append(pl_data_dict)
                                else:
                                    for ent_member in notch_list:
                                        if ent_member['A'] == rgb_str and ent_member['B'] == width and \
                                                ent_member['C'] == height:
                                            ent_member['D'] += 1
                                            found = True
                                            break
                                    if not found:
                                        pl_data_dict = self.__compose_data_dict(rgb_str, width, height, count)
                                        notch_list.append(pl_data_dict)
                        block_name = block_ent.dxf.name
                        # for every block you create 1 worksheet
                        self.__create_spreadsheet(workbook, block_name, pl_list, notch_list, cad_manager.unit)
                    filename_ext = os.path.basename(file_full_path)
                    filename = os.path.splitext(filename_ext)[0]
                    output_dir = os.path.join(Constants.OUTPUT_DIR, filename + '.xlsx')
                    if len(workbook.worksheets) > 1:
                        default_ws = workbook["Sheet"]
                        workbook.remove(default_ws)
                        workbook.save(output_dir)
                    else:
                        print(f"No BlockReference Found inside file: {file_full_path}")

    def __is_notch(self, polyline):
        is_notch = False

        ent_point = polyline.get_points('xy')
        rounded_ent_points = [(round(point[0], Constants.ROUND_PRECISION), round(point[1], Constants.ROUND_PRECISION))
                              for point in ent_point]

        bbox = ezdxf.math.BoundingBox2d(rounded_ent_points)
        extmin = bbox.extmin
        extmax = bbox.extmax

        # getting the extremas of the bounding box
        xmax, ymax = extmax
        xmin, ymin = extmin

        four_corners = [(xmin, ymax), (xmax, ymax), (xmin, ymin), (xmax, ymin)]

        vec_pl_points = self.__convert_to_vec2s(rounded_ent_points)
        for corner_point in four_corners:
            vec_corn_point = ezdxf.math.Vec2(corner_point)
            if ezdxf.math.is_point_in_polygon_2d(vec_corn_point, vec_pl_points, abs_tol=1e-5) == -1:
                is_notch = True
                break
        return is_notch

    def __convert_to_vec2s(self, pt_list):
        vec2_list = []
        for pt in pt_list:
            vec2_pt = ezdxf.math.Vec2(pt)
            vec2_list.append(vec2_pt)
        return vec2_list

    def __create_spreadsheet(self, workbook, block_name, ent_list, notch_list, unit):
        # add unit conversions that are necessary
        new_unit = unit
        if unit == "Inches":
            new_unit = "Feet-Inch"
        fieldnames = ['RGBColor', f'Width ({new_unit})', f'Height ({new_unit})', 'Count']
        worksheet = workbook.create_sheet(f"BlockReference {block_name}")
        worksheet.append(fieldnames)
        for i, ent_prop in enumerate(ent_list):
            worksheet.append(ent_prop)
            color = Color(rgb=ent_prop['A'])
            fill = PatternFill(patternType='solid', fgColor=color)
            worksheet[f'A{i+2}'].fill = fill

        worksheet.append([""])
        worksheet.append(["Notches:"])
        for i, ent_prop in enumerate(notch_list):
            worksheet.append(ent_prop)
            color = Color(rgb=ent_prop['A'])
            fill = PatternFill(patternType='solid', fgColor=color)
            worksheet[f'A{i+2}'].fill = fill

    def __compose_data_dict(self, rgb_color, width, height, count):
        ent_prop_dict = {
            'A': rgb_color,
            'B': width,
            'C': height,
            'D': count
        }
        return ent_prop_dict

    def __get_rgb(self, polyline):
        # getting the autocad color index of the polyline
        ent_aci = polyline.dxf.color
        try:
            ent_color_hex = DXF_DEFAULT_COLORS[ent_aci]
        except IndexError:
            try:
                ent_color_hex = DXF_DEFAULT_COLORS_2[ent_aci]
            except IndexError:
                ent_color_hex = None
        # if ent_color_hex:
        #     ent_color_rgb = int2rgb(ent_color_hex)
        #     ent_color_rgb_str = f"({ent_color_rgb[0]}, {ent_color_rgb[1]}, {ent_color_rgb[2]})"
        # else:
        #     ent_color_rgb_str = "Color Undefined"
        try:
            ent_color_hex_str = f"{ent_color_hex:06X}"
        except TypeError:
            ent_color_hex_str = "FFFFFF"
        return ent_color_hex_str

    def __compute_width_height(self, polyline, unit):
        # getting the points for creation of bounding box
        ent_points = polyline.get_points('xy')
        bbox = ezdxf.math.BoundingBox2d(ent_points)
        extmin = bbox.extmin
        extmax = bbox.extmax

        # getting the extremas of the bounding box
        xmax, ymax = extmax
        xmin, ymin = extmin

        # getting the width and the height of the polyline
        w = round(xmax - xmin, 2)
        h = round(ymax - ymin, 2)

        # add unit conversions that are necessary
        # convert inches to feet-inch
        if unit == "Inches":
            w = self.__convert_inches_to_feet_inch(w)
            h = self.__convert_inches_to_feet_inch(h)
        return w, h

    def __convert_inches_to_feet_inch(self, inches):
        # convert inches to feet
        feet = round(inches / 12, 2)
        # get the whole part and decimal part
        decimal_feet, whole_feet = math.modf(feet)
        whole_feet = int(whole_feet)
        decimal_feet = round(decimal_feet, 2)
        # convert the decimal part to inches
        new_inches = round(decimal_feet * 12, 2)

        feet_inches_fraction = f"{whole_feet}' - {new_inches}\""
        return feet_inches_fraction
