import datetime
from src.business_logic.calculator import Calculator
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.cell.cell import Cell


class EmployeeSpreadsheet:
    def __init__(self, workbook, employee, tasks, holidays):
        self.ts_workbook = workbook
        self.employee = employee
        self.tasks = tasks
        self.holidays = holidays
        self.employee_full_name = "{}, {} {}.".format(employee.get_last_name(),
                                                      employee.get_first_name(),
                                                      employee.get_middle_name()[0])
        self.ws = self.ts_workbook.workbook.create_sheet(self.employee_full_name)

    def write_headers(self):
        self.ws["A1"] = "MANHOUR ANALYSIS OF RENDERED DUTIES"
        self.ws["A2"] = "PERIOD COVERED: {} to {}".format(self.ts_workbook.get_start_date().strftime("%d-%b-%Y"),
                                                          self.ts_workbook.get_end_date().strftime("%d-%b-%Y"))
        self.ws["A3"] = self.employee_full_name
        self.ws["A3"].font = Font(size=12, bold=True, name='Calibri')

    def write_body(self):
        days_dict = {
            0: "MON",
            1: "TUE",
            2: "WED",
            3: "THU",
            4: "FRI",
            5: "SAT",
            6: "SUN"
        }
        # get the weekly cluster in dates
        weeks = self.__generate_mon_to_sun_dates()
        for week in weeks:
            first_day = week[0]
            last_day = week[-1]
            self.ws.append([""])
            self.ws.append(self.__apply_bold_cells(["WEEK COVERED: {} to {}".format(first_day.strftime("%d-%b-%Y"),
                                                                                    last_day.strftime("%d-%b-%Y"))]))
            self.__write_employee_report_columns(self.ws)
            for day in week:
                # if holiday, treat all hours as excess
                if day in self.holidays:
                    credited_minutes = 0
                    date_string = day.strftime("%d-%b-%Y") + " (HOLIDAY)"
                else:
                    credited_minutes = 480
                    date_string = day.strftime("%d-%b-%Y")
                day_name = days_dict[day.weekday()]
                tito = self.__get_tito_of_the_day(day, self.tasks)
                total_mins = self.__get_total_minutes_of_the_day(day, self.tasks)
                excess_minutes = 0
                if total_mins > credited_minutes:
                    excess_minutes = total_mins - credited_minutes
                if tito != "":
                    data = [date_string, day_name, tito, total_mins, credited_minutes, excess_minutes]
                else:
                    data = [date_string, day_name, tito, total_mins, 0, excess_minutes]
                self.ws.append(self.__apply_border_cells(data))

    def __apply_border_cells(self, data):
        for value in data:
            c = Cell(self.ws, column="A", row=1, value=value)
            thin_border = Border(left=Side(style="thin"),
                                 right=Side(style="thin"),
                                 top=Side(style="thin"),
                                 bottom=Side(style="thin"))
            c.border = thin_border
            yield c

    def __apply_bold_cells(self, data):
        for value in data:
            c = Cell(self.ws, column="A", row=1, value=value)
            font = Font(bold=True, name='Calibri')
            c.font = font
            yield c

    def __get_tito_of_the_day(self, date, tasks):
        tito_list = []
        for task in tasks:
            task_date = task.get_date()
            if task_date == date:
                tito_list.append("{}-{}".format(task.get_time_in().strftime("%I:%M %p"),
                                                task.get_time_out().strftime("%I:%M %p")))
        tito = "\n".join(tito_list)
        return tito

    def __get_total_minutes_of_the_day(self, date, tasks):
        calculator = Calculator()
        total_minutes = 0
        for task in tasks:
            task_date = task.get_date()
            if task_date == date:
                task_hours = calculator.calculate_task_hours(task_date, task.get_time_in(), task.get_time_out())
                total_minutes += task_hours * 60  # hours to minutes
        return total_minutes

    def __write_employee_report_columns(self, ws):
        data = ["Date", "Day", "Time-in/Time-out", "Rendered MINS for the Day", "Credited Regular Log [480 = 1 day]",
                "Minutes in excess of 480; Sat/Sun Duties"]
        ws.append(self.__apply_border_cells(data))

    def __generate_mon_to_sun_dates(self):
        week = []
        weeks = []
        start_date = self.ts_workbook.get_start_date()
        end_date = self.ts_workbook.get_end_date()

        dates = [start_date + datetime.timedelta(days=x) for x in range((end_date - start_date).days + 1)]

        for i, date in enumerate(dates):
            week.append(date)
            if date.weekday() == 6 or (date.weekday() != 6 and i == len(dates) - 1):
                weeks.append(week)
                week = []
        return weeks
