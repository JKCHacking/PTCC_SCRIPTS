import datetime
from src.business_logic.calculator import Calculator
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.cell.cell import Cell


class SummarySpreadsheet:
    def __init__(self, workbook):
        self.ts_workbook = workbook
        self.ws = self.ts_workbook.workbook.create_sheet("Summary")

    def write_headers(self):
        self.ws["A1"] = "PTCC TIMESHEET FOR WEEK ({} to {})".format(
            self.ts_workbook.get_start_date().strftime("%d-%b-%Y"),
            self.ts_workbook.get_end_date().strftime("%d-%b-%Y"))
        self.ws["A2"] = datetime.datetime.now().strftime("%d-%B-%Y")
        self.ws.append([""])
        self.ws.append(self.__apply_border_cells(["NAME", "PROJECT", "TIME SPENT"]))

    def write_body(self, employee, tasks):
        name = "{}, {} {}.".format(employee.get_last_name(),employee.get_first_name(),employee.get_middle_name()[0])
        projects = self.__get_projects(tasks)
        projects = "\n".join(sorted(projects))
        time_spent = self.__get_total_hours(tasks)
        self.ws.append(self.__apply_border_cells([name, projects, time_spent]))

    def write_footers(self):
        self.ws.append([""])
        self.ws.append(["APPROVED by (SUPERVISOR)"])
        self.ws.append([""])
        self.ws.append(["_________________________"])
        self.ws.append(["ADAM LEE"])
        self.ws.append([""])
        self.ws.append([""])
        self.ws.append(["NOTED BY:"])
        self.ws.append([""])
        self.ws.append(["_________________________"])
        self.ws.append(["SUSAN BENJAMIN"])

    def __get_projects(self, tasks):
        project_set = set()
        for task in tasks:
            project_set.add(task.get_project_name())
        return project_set

    def __get_total_hours(self, tasks):
        calculator = Calculator()
        total_hours = 0
        for task in tasks:
            if self.ts_workbook.get_start_date() <= task.get_date() <= self.ts_workbook.get_end_date():
                total_hours += calculator.calculate_task_hours(task.get_date(), task.get_time_in(), task.get_time_out())
        return total_hours

    def __apply_border_cells(self, data):
        for value in data:
            c = Cell(self.ws, column="A", row=1, value=value)
            thin_border = Border(left=Side(style="thin"),
                                 right=Side(style="thin"),
                                 top=Side(style="thin"),
                                 bottom=Side(style="thin"))
            c.border = thin_border
            yield c

    def __apply_bold_cells(self, data):
        for value in data:
            c = Cell(self.ws, column="A", row=1, value=value)
            font = Font(bold=True, name='Calibri')
            c.font = font
            yield c
