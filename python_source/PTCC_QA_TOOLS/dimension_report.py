import openpyxl
import math
import tkinter
from comtypes import client
from tkinter.filedialog import askopenfilenames
from min_bb import get_3d_bb


class DimensionReport:
    @staticmethod
    def get_dimensions(min_pt, max_pt):
        width = max_pt[0] - min_pt[0]
        length = max_pt[1] - min_pt[1]
        height = max_pt[2] - max_pt[2]
        return length, width, height

    def create_report(self, doc, objs, out_file_name):
        groups = {}
        ac_color_method_by_layer = 192
        ac_color_method_by_aci = 195
        ac_color_method_by_rgb = 194

        wb = openpyxl.Workbook()
        ws_obj_detail = wb.create_sheet("Dimension Details")
        ws_obj_count = wb.create_sheet("Object Count")

        ws_obj_detail.cell(row=1, column=1, value="Handle")
        ws_obj_detail.cell(row=1, column=2, value="Length")
        ws_obj_detail.cell(row=1, column=3, value="Width")
        ws_obj_detail.cell(row=1, column=4, value="Height")
        ws_obj_detail.cell(row=1, column=5, value="Color")
        for i, obj in enumerate(objs):
            min_pt, max_pt = get_3d_bb(doc, obj, math.radians(2))
            length, width, height = self.get_dimensions(min_pt, max_pt)
            handle = obj.Handle
            color = 0
            if obj.TrueColor.ColorMethod == ac_color_method_by_layer:
                color = doc.Layers[obj.Layer].TrueColor.ColorIndex
            elif (obj.TrueColor.ColorMethod == ac_color_method_by_aci or
                  obj.TrueColor.ColorMethod == ac_color_method_by_rgb):
                color = obj.TrueColor.ColorIndex
            ws_obj_detail.cell(row=i + 2, column=1, value=handle)
            ws_obj_detail.cell(row=i + 2, column=2, value=length)
            ws_obj_detail.cell(row=i + 2, column=3, value=width)
            ws_obj_detail.cell(row=i + 2, column=4, value=height)
            ws_obj_detail.cell(row=i + 2, column=5, value=color)

            if length != 0:
                if length in groups:
                    groups.update({length: groups[length] + 1})
                else:
                    groups.update({length: 1})
            else:
                print("solid with 0 length: {}".format(obj.Handle))

        ws_obj_count.cell(row=1, column=1, value="Length")
        ws_obj_count.cell(row=1, column=2, value="Count")
        for i, key, value in enumerate(groups.items()):
            ws_obj_count.cell(row=i + 2, column=1, value=key)
            ws_obj_count.cell(row=i + 2, column=2, value=value)
        wb.save(out_file_name)


def get_objects(doc, obj_name):
    objs = []
    for obj in doc.ModelSpace:
        if obj.ObjectName.lower() == obj_name:
            objs.append(obj)
    return objs


if __name__ == "__main__":
    tkinter.Tk().withdraw()
    cad_app = client.GetActiveObject("BricscadApp.AcadApplication")
    dwg_file_paths = askopenfilenames(title="Select multiple DWG", filetypes=[("DWG Files", ".dwg")])
    dim_report = DimensionReport()
    if dwg_file_paths:
        for dwg_file_path in dwg_file_paths:
            doc = cad_app.Documents.Open(dwg_file_path)
            objs = get_objects(doc, "acdb3dsolid")
            dim_report.create_report(doc, objs, "Dimension Report.xlsx")
