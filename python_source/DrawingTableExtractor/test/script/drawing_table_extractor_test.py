import os
import unittest
import time
from openpyxl import load_workbook
from src.script.drawing_table_extractor import main
from src.util.constants import Constants


class DrawingTableExtractorTest(unittest.TestCase):

    def __common(self, test_filepath, expected_filepath):
        expected_sheetnames = ["W.01.DL.08A", "W.01.DL.08B", "W.01.DL.08C", "W.01.DL.08D", "W.01.DL.08E"]

        pdf1 = os.path.join(os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "W.01.DL.08A.pdf"))
        pdf2 = os.path.join(os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "W.01.DL.08B.pdf"))
        pdf3 = os.path.join(os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "W.01.DL.08C.pdf"))
        pdf4 = os.path.join(os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "W.01.DL.08D.pdf"))
        pdf5 = os.path.join(os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "W.01.DL.08E.pdf"))

        main(test_filepath)

        self.assertTrue(expected_filepath)
        workbook = load_workbook(expected_filepath)
        sheetnames = workbook.sheetnames
        for actual, expected in zip(sheetnames, expected_sheetnames):
            self.assertEqual(actual, expected)

        worksheets = workbook.worksheets
        self.assertEqual(worksheets[0].cell(row=3, column=1).value, "1")
        self.assertEqual(worksheets[0].cell(row=3, column=2).value, "01.COVER")
        self.assertEqual(worksheets[0].cell(row=3, column=3).value, "COVER PAGE")
        self.assertEqual(worksheets[0].cell(row=3, column=4).value, "-")
        self.assertEqual(worksheets[0].cell(row=3, column=5).value, "21 DECEMBER 2018")
        self.assertEqual(worksheets[0].cell(row=3, column=6).value, "A")
        self.assertEqual(worksheets[0].cell(row=3, column=7).value, "28 DECEMBER 2018")
        self.assertEqual(worksheets[0].cell(row=3, column=8).value, "B")
        self.assertEqual(worksheets[0].cell(row=3, column=9).value, "11 JANUARY 2019")
        self.assertEqual(worksheets[0].cell(row=3, column=10).value, "C")
        self.assertEqual(worksheets[0].cell(row=3, column=11).value, "18 JANUARY 2019")
        self.assertFalse(os.path.exists(pdf1))
        self.assertFalse(os.path.exists(pdf2))
        self.assertFalse(os.path.exists(pdf3))
        self.assertFalse(os.path.exists(pdf4))
        self.assertFalse(os.path.exists(pdf5))
        
        os.remove(expected_filepath)

    def test_main_file(self):
        # expected:
        #     1. 1 excel file with drawing_name as filename exists.
        #     2. values inside each worksheet should be consistent with the pdf files respectively.
        #     3. worksheet names should be equal to the layout names
        #     4. Bricscad application should be closed.
        #     5. there should be no pdf files with layout names as filenames

        test_filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "testdata1.dwg")
        expected_filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "testdata1.xlsx")
        self.__common(test_filepath, expected_filepath)

    def test_main_folder(self):
        # expected:
        #     1. 1 excel file with drawing_name as filename exists.
        #     2. values inside each worksheet should be consistent with the pdf files respectively.
        #     3. worksheet names should be equal to the layout names
        #     4. Bricscad application should be closed.
        #     5. there should be no pdf files with layout names as filenames
        
        test_filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_dir")
        expected_filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "testdata1.xlsx")
        self.__common(test_filepath, expected_filepath)
