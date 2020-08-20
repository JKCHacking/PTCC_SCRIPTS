import unittest
import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from src.document.exceldocument import ExcelDocument
from src.util.constants import Constants


class ExcelDocumentTest(unittest.TestCase):

    def test_save(self):
        filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_excel_doc.xlsx")
        workbook = Workbook()
        excel_doc = ExcelDocument(workbook)
        excel_doc.save(filepath)

        self.assertTrue(os.path.exists(filepath))
        os.remove(filepath)

    def test_create_worksheet(self):
        sheet_name = "test"
        filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_excel_doc.xlsx")
        workbook = Workbook()
        excel_doc = ExcelDocument(workbook)
        excel_doc.create_worksheet(sheet_name)
        excel_doc.save(filepath)

        excel_doc = load_workbook(filepath)
        actual = excel_doc.sheetnames[1]
        expected = sheet_name
        self.assertEqual(actual, expected)
        os.remove(filepath)

    def test_add_worksheet_contents(self):
        sheet_name = "test"
        test_content = "test_content"
        filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_excel_doc.xlsx")
        workbook = Workbook()
        excel_doc = ExcelDocument(workbook)
        excel_worksheet = excel_doc.create_worksheet(sheet_name)
        excel_doc.add_worksheet_contents(excel_worksheet, test_content, (1, 1))
        excel_doc.save(filepath)

        excel_doc = load_workbook(filepath)
        worksheet = excel_doc[sheet_name]
        actual = worksheet.cell(row=1, column=1).value
        expected = test_content
        self.assertEqual(actual, expected)
        os.remove(filepath)
