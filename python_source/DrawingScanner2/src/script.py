import os
import math
import ezdxf
import ezdxf.math
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from ezdxf.tools.rgb import DXF_DEFAULT_COLORS, DXF_DEFAULT_COLORS_2, int2rgb
from src.util.constants import Constants


class Script:
    """
        Gets all polylines from a dxf file and records the handle, width, and height of each polyline.
    """
    def __init__(self):
        self.unit = None

    def iter_input(self):
        """
            gets every file in the output folder and process the necessary data
        """
        input_dir = Constants.INPUT_DIR
        for dir_path, dir_names, file_names in os.walk(input_dir):
            for file_name in file_names:
                file_full_path = os.path.join(dir_path, file_name)
                if file_full_path.endswith(Constants.DXF_FILE_EXT):
                    # for every dxf file you create 1 excel file
                    workbook = Workbook()
                    # getting all the blocks
                    for block_ent in self.__get_blocks(file_full_path):
                        # polyline container
                        pl_list = []
                        # getting all the polylines in the block reference
                        for entity in block_ent.virtual_entities():
                            type = entity.dxftype()
                            if type == "LWPOLYLINE" or type == "POLYLINE":
                                found = False
                                count = 1
                                width, height = self.__compute_width_height(entity)
                                rgb_str = self.__get_rgb(entity)
                                # checking if same polyline already exist in the list.
                                for ent_member in pl_list:
                                    if ent_member['A'] == rgb_str and ent_member['C'] == height:
                                        ent_member['D'] += 1
                                        found = True
                                        break
                                if not found:
                                    pl_data_dict = self.__compose_data_dict(rgb_str, width, height, count)
                                    pl_list.append(pl_data_dict)
                        block_name = block_ent.dxf.name
                        # for every block you create 1 worksheet
                        self.__create_spreadsheet(workbook, block_name, pl_list)
                    filename_ext = os.path.basename(file_full_path)
                    filename = os.path.splitext(filename_ext)[0]
                    output_dir = os.path.join(Constants.OUTPUT_DIR, filename + '.xlsx')
                    if len(workbook.worksheets) > 1:
                        default_ws = workbook["Sheet"]
                        workbook.remove(default_ws)
                        workbook.save(output_dir)
                    else:
                        print(f"No BlockReference Found inside file: {file_full_path}")

    def __create_spreadsheet(self, workbook, block_name, ent_list):
        # add unit conversions that are necessary
        new_unit = self.unit
        if self.unit == "Inches":
            new_unit = "Feet-Inch"
        fieldnames = ['RGBColor', f'Width ({new_unit})', f'Height ({new_unit})', 'Count']
        worksheet = workbook.create_sheet(f"BlockReference {block_name}")
        worksheet.append(fieldnames)
        for i, ent_prop in enumerate(ent_list):
            worksheet.append(ent_prop)
            color = Color(rgb=ent_prop['A'])
            fill = PatternFill(patternType='solid', fgColor=color)
            worksheet[f'A{i+2}'].fill = fill

    def __compose_data_dict(self, rgb_color, width, height, count):
        ent_prop_dict = {
            'A': rgb_color,
            'B': width,
            'C': height,
            'D': count
        }
        return ent_prop_dict

    def __get_rgb(self, polyline):
        # getting the autocad color index of the polyline
        ent_aci = polyline.dxf.color
        try:
            ent_color_hex = DXF_DEFAULT_COLORS[ent_aci]
        except IndexError:
            try:
                ent_color_hex = DXF_DEFAULT_COLORS_2[ent_aci]
            except IndexError:
                ent_color_hex = None
        # if ent_color_hex:
        #     ent_color_rgb = int2rgb(ent_color_hex)
        #     ent_color_rgb_str = f"({ent_color_rgb[0]}, {ent_color_rgb[1]}, {ent_color_rgb[2]})"
        # else:
        #     ent_color_rgb_str = "Color Undefined"
        try:
            ent_color_hex_str = f"{ent_color_hex:06X}"
        except TypeError:
            ent_color_hex_str = "FFFFFF"
        return ent_color_hex_str

    def __compute_width_height(self, polyline):
        # getting the points for creation of bounding box
        ent_points = polyline.get_points('xy')
        bbox = ezdxf.math.BoundingBox2d(ent_points)
        extmin = bbox.extmin
        extmax = bbox.extmax

        # getting the extremas of the bounding box
        xmax, ymax = extmax
        xmin, ymin = extmin

        # getting the width and the height of the polyline
        w = round(xmax - xmin, 2)
        h = round(ymax - ymin, 2)

        # add unit conversions that are necessary
        # convert inches to feet-inch
        if self.unit == "Inches":
            w = self.__convert_inches_to_feet_inch(w)
            h = self.__convert_inches_to_feet_inch(h)
        return w, h

    def __get_blocks(self, dxf_fp):
        dxf_file = ezdxf.readfile(dxf_fp)
        ins_unit_num = dxf_file.header["$INSUNITS"]
        self.unit = Constants.INS_UNITS[ins_unit_num]
        ms = dxf_file.modelspace()
        for ent in ms:
            if ent.dxftype() == "INSERT":
                block_ent = ent
                yield block_ent

    def __convert_inches_to_feet_inch(self, inches):
        # convert inches to feet
        feet = round(inches / 12, 2)
        # get the whole part and decimal part
        decimal_feet, whole_feet = math.modf(feet)
        whole_feet = int(whole_feet)
        decimal_feet = round(decimal_feet, 2)
        # convert the decimal part to inches
        new_inches = round(decimal_feet * 12, 2)

        feet_inches_fraction = f"{whole_feet}' - {new_inches}\""
        return feet_inches_fraction
