import os
import datetime
import openpyxl
from src.utils.constants import Constants
from src.entity.employee import Employee
from src.entity.task import Task
from src.entity.employee_timesheet import EmployeeTimesheet


class TimesheetParser:
    def __init__(self):
        pass

    def iter_spreadsheets(self):
        directory = Constants.INPUT_DIR
        for dirpath, dirnames, filenames in os.walk(directory):
            for filename in filenames:
                fullpath = os.path.join(dirpath, filename)
                if fullpath.endswith(Constants.EXCEL_EXT):
                    yield fullpath

    def parse(self):
        timesheet_list = []
        for excel_fp in self.iter_spreadsheets():
            wb = openpyxl.load_workbook(excel_fp)
            ws = wb[wb.sheetnames[0]]
            id_number = ws["A6"].value
            first_name = ws["B6"].value
            middle_name = ws["C6"].value
            last_name = ws["D6"].value
            # create Employee object
            emp = Employee(id_number=id_number,
                           first_name=first_name,
                           middle_name=middle_name,
                           last_name=last_name)
            # create list of Task objects
            task_list = []
            row_num = 11
            while not all(cell.value is None for row in ws["A{}".format(row_num):"E{}".format(row_num)]
                          for cell in row):
                date = ws["A{}".format(row_num)].value
                start_time = ws["B{}".format(row_num)].value
                end_time = ws["C{}".format(row_num)].value
                task_name = ws["D{}".format(row_num)].value
                project_name = ws["E{}".format(row_num)].value
                # date, start_time and end_time must be valid
                # to avoid error in the calculator later
                if isinstance(date, datetime.datetime):
                    date = date.date()
                if isinstance(start_time, datetime.datetime):
                    start_time = start_time.time()
                if isinstance(end_time, datetime.datetime):
                    end_time = end_time.time()
                if isinstance(date, datetime.date) and isinstance(start_time, datetime.time) and \
                        isinstance(end_time, datetime.time):
                    task = Task(date=date,
                                start_time=start_time,
                                end_time=end_time,
                                task_name=task_name,
                                project_name=project_name)
                    task_list.append(task)
                row_num += 1
            emp_ts = EmployeeTimesheet(
                employee=emp,
                task_list=task_list
            )
            timesheet_list.append(emp_ts)
        return timesheet_list
