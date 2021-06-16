import os
import openpyxl
from openpyxl.utils import get_column_letter
from src.utils.constants import Constants


class TimesheetWorkbook:
    def __init__(self, start_date, end_date):
        self.start_date = start_date
        self.end_date = end_date
        self.workbook = openpyxl.Workbook()
        self.output_path = os.path.join(Constants.OUTPUT_DIR, "Employee Timesheet {} to {}.xlsx".format(
            self.start_date.strftime("%d-%b-%Y"), self.end_date.strftime("%d-%b-%Y")))

    def get_start_date(self):
        return self.start_date

    def get_end_date(self):
        return self.end_date

    def get_output_path(self):
        return self.output_path

    def get_workbook(self):
        return self.workbook

    def adjust_cell_height_width(self):
        for sheetname in self.workbook.sheetnames:
            ws = self.workbook[sheetname]
            dims_col = {}
            dims_row = {}

            for i, row in enumerate(ws.rows):
                for cell in row:
                    cell_value = cell.value
                    if cell_value:
                        dims_col[cell.column_letter] = max((dims_col.get(cell.column_letter, 0),
                                                            len(str(cell.value)))) + 1.2
                        newline_count = str(cell_value).count('\n')
                        max_value = max((dims_row.get(i + 1, 0), newline_count))
                        dims_row[i + 1] = max_value * 4.7

            for col, value in dims_col.items():
                ws.column_dimensions[col].width = value

            for row, value in dims_row.items():
                if value != 0:
                    ws.row_dimensions[row].height = value

    def __sort_employee_sheets(self):
        summary_sheet = self.workbook._sheets[0]
        employee_sheets = self.workbook._sheets[1:]
        # sort by name
        employee_sheets.sort(key=lambda ws: ws.title)
        self.workbook._sheets = [summary_sheet] + employee_sheets

    def save(self):
        self.workbook.remove(self.workbook["Sheet"])
        self.adjust_cell_height_width()
        self.__sort_employee_sheets()
        self.workbook.save(self.output_path)
