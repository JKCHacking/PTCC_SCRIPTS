import unittest
import unittest
import os
import datetime
import openpyxl
from src.business_logic.timesheet_workbook import TimesheetWorkbook
from src.business_logic.employee_spreadsheet import EmployeeSpreadsheet
from src.entity.employee import Employee
from src.entity.task import Task


class EmployeeSpreadsheetTest(unittest.TestCase):
    def test_generate_employee_spreadsheet(self):
        employee_name = "Doe, Jean S."
        expected_wb_fpath = "testdata/ss_generator_testdata_001.xlsx"
        expected_wb = openpyxl.load_workbook(expected_wb_fpath)
        expected_ws = expected_wb[employee_name]

        s_date = datetime.date(day=7, month=6, year=2021)
        e_date = datetime.date(day=21, month=6, year=2021)

        emp = Employee("1234", "Jean", "Spring", "Doe")
        task_list = [Task(datetime.date(day=7, month=6, year=2021), time_in=datetime.time(hour=9),
                          time_out=datetime.time(hour=14), project_name="Project1", task_name="Task1"),
                     Task(datetime.date(day=7, month=6, year=2021), time_in=datetime.time(hour=14),
                          time_out=datetime.time(hour=19), project_name="Project2", task_name="Task2"),
                     Task(datetime.date(day=8, month=6, year=2021), time_in=datetime.time(hour=9),
                          time_out=datetime.time(hour=18), project_name="Project3", task_name="Task3"),
                     Task(datetime.date(day=9, month=6, year=2021), time_in=datetime.time(hour=9),
                          time_out=datetime.time(hour=18), project_name="Project4", task_name="Task4"),
                     Task(datetime.date(day=10, month=6, year=2021), time_in=datetime.time(hour=9),
                          time_out=datetime.time(hour=18), project_name="Project5", task_name="Task5"),
                     Task(datetime.date(day=11, month=6, year=2021), time_in=datetime.time(hour=9),
                          time_out=datetime.time(hour=18), project_name="Project6", task_name="Task6"),
                     Task(datetime.date(day=12, month=6, year=2021), time_in=datetime.time(hour=9),
                          time_out=datetime.time(hour=18), project_name="Project7", task_name="Task7"),
                     Task(datetime.date(day=14, month=6, year=2021), time_in=datetime.time(hour=9),
                          time_out=datetime.time(hour=18), project_name="Project8", task_name="Task8"),
                     Task(datetime.date(day=15, month=6, year=2021), time_in=datetime.time(hour=9),
                          time_out=datetime.time(hour=18), project_name="Project9", task_name="Task9"),
                     Task(datetime.date(day=19, month=6, year=2021), time_in=datetime.time(hour=17),
                          time_out=datetime.time(hour=20), project_name="Project10", task_name="Task10")
                     ]
        holiday_list = [datetime.date(day=12, month=6, year=2021)]
        ts_wb = TimesheetWorkbook(s_date, e_date)
        emp_ss = EmployeeSpreadsheet(ts_wb, emp, task_list, holiday_list)
        emp_ss.write_headers()
        emp_ss.write_body()
        ts_wb.save()

        actual_wb_fpath = ts_wb.get_output_path()
        actual_wb = openpyxl.load_workbook(actual_wb_fpath)
        self.assertTrue(employee_name in actual_wb.sheetnames)

        # compare all expected cells to actual cells
        actual_ws = actual_wb[employee_name]
        for expected_col, actual_col in zip(expected_ws.iter_cols(), actual_ws.iter_cols()):
            for expected_cell, actual_cell in zip(expected_col, actual_col):
                if expected_cell.value is None:
                    self.assertIsNone(actual_cell.value)
                else:
                    self.assertEqual(expected_cell.value, actual_cell.value)

        os.remove(actual_wb_fpath)
