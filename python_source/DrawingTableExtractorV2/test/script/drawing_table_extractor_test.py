import os
import unittest
import time
from comtypes import client, COMError
from openpyxl import load_workbook
from src.script.drawing_table_extractor_v2 import is_inside_box, get_entities_within_box, main, is_between_line, get_entities_between_points
from src.util.constants import Constants
from src.util.util import Utilities


class DrawingTableExtractorTest(unittest.TestCase):
    def setUp(self) -> None:
        try:
            self.cad_application = client.GetActiveObject(Constants.BRICS_APP_NAME, dynamic=True)
            self.cad_application.Visible = False
        except(OSError, COMError):
            self.cad_application = client.CreateObject(Constants.BRICS_APP_NAME, dynamic=True)
            self.cad_application.Visible = False
        return self.cad_application

    def tearDown(self) -> None:
        self.cad_application.Quit()

    def test_is_inside_box(self):
        min = (0, 0)
        max = (5, 5)
        point = (2, 3)
        self.assertTrue(is_inside_box(min, max, point))

    def test_is_outside_box(self):
        min = (0, 0)
        max = (5, 5)
        point = (3, 6)
        self.assertFalse(is_inside_box(min, max, point))
    
    def test_outside_flat_box(self):
        min = (0, 0)
        max = (0, 5)
        point = (2.5, 2.5)
        self.assertFalse(is_inside_box(min, max, point))

    def test_inside_flat_box(self):
        min = (0, 0)
        max = (0, 5)
        point = (0, 1)
        self.assertTrue(is_inside_box(min, max, point))

    def test_is_between_v_line(self):
        left = (0, 0)
        right = (5, 0)
        point1 = (2.5, -2.5)
        point2 = (2.5, 2.5)
        self.assertTrue(is_between_line(left, right, point1))
        self.assertTrue(is_between_line(left, right, point2))

    def test_is_outside_v_line(self):
        left = (0, 0)
        right = (5, 0)
        point1 = (10, 5)
        point2 = (-10, 5)
        self.assertFalse(is_between_line(left, right, point1))
        self.assertFalse(is_between_line(left, right, point2))

    def test_is_between_h_line(self):
        left = (0, 0)
        right = (0, 5)
        point1 = (2.5, 2.5)
        point2 = (-2.5, 2.5)
        self.assertTrue(is_between_line(left, right, point1))
        self.assertTrue(is_between_line(left, right, point2))

    def test_is_outside_h_line(self):
        left = (0, 0)
        right = (0, 5)
        point1 = (10, 10)
        point2 = (-10, -10)
        self.assertFalse(is_between_line(left, right, point1))
        self.assertFalse(is_between_line(left, right, point2))

    def test_get_entities_within_box(self):
        testdata = os.path.join(Constants.TEST_DIR, "testdata", "testdata_boundary.dwg")
        doc = self.cad_application.Documents.Open(testdata)
        modelspace = doc.ModelSpace

        expected = ["01.COVER", "COVER PAGE", "WEST F2 - TOP HAT GENERAL ASSEMBLY"]
        ent_ls = get_entities_within_box(modelspace, ['AcDbMText', 'AcDbText'], (0, 0, 0), (5, 5, 0))
        self.assertEqual(3, len(ent_ls))
        for ent in ent_ls:
            self.assertTrue(ent.TextString in expected)

    def test_get_entities_between_line(self):
        testdata = os.path.join(Constants.TEST_DIR, "testdata", "testdata_line_boundary.dwg")
        doc = self.cad_application.Documents.Open(testdata)
        modelspace = doc.ModelSpace

        expected = ["01.COVER", "COVER PAGE", "WEST F2 - TOP HAT GENERAL ASSEMBLY"]
        ent_ls = get_entities_between_points(modelspace, ['AcDbMText', 'AcDbText'], (0, 10, 0), (5, 10, 0))
        self.assertEqual(3, len(ent_ls))
        for ent in ent_ls:
            self.assertTrue(ent.TextString in expected)

    def __common(self, test_filepath, expected_filepath):
        expected_sheetnames = ["testdata1A", "testdata1B", "testdata1C", "testdata1D", "testdata1E"]
        main(test_filepath)

        self.assertTrue(expected_filepath)
        workbook = load_workbook(expected_filepath)
        sheetnames = workbook.sheetnames
        for actual, expected in zip(sheetnames, expected_sheetnames):
            self.assertEqual(actual, expected)

        worksheets = workbook.worksheets
        self.assertEqual(worksheets[0].cell(row=3, column=1).value, "1")
        self.assertEqual(worksheets[0].cell(row=3, column=2).value, "01.COVER")
        self.assertEqual(worksheets[0].cell(row=3, column=3).value, "COVER PAGE")
        self.assertEqual(worksheets[0].cell(row=3, column=4).value, "-")
        self.assertEqual(worksheets[0].cell(row=3, column=5).value, "21 DECEMBER 2018")
        self.assertEqual(worksheets[0].cell(row=3, column=6).value, "A")
        self.assertEqual(worksheets[0].cell(row=3, column=7).value, "28 DECEMBER 2018")
        self.assertEqual(worksheets[0].cell(row=3, column=8).value, "B")
        self.assertEqual(worksheets[0].cell(row=3, column=9).value, "11 JANUARY 2019")
        self.assertEqual(worksheets[0].cell(row=3, column=10).value, "C")
        self.assertEqual(worksheets[0].cell(row=3, column=11).value, "18 JANUARY 2019")

        os.remove(expected_filepath)

    def test_main_file(self):
        # expected:
        #     1. 1 excel file with drawing_name as filename exists.
        #     2. values inside each worksheet should be consistent with the pdf files respectively.
        #     3. worksheet names should be equal to the layout names
        #     4. Bricscad application should be closed.
        #     5. there should be no pdf files with layout names as filenames

        test_filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "testdata1.dwg")
        expected_filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "testdata1.xlsx")
        self.__common(test_filepath, expected_filepath)

    def test_main_folder(self):
        # expected:
        #     1. 1 excel file with drawing_name as filename exists.
        #     2. values inside each worksheet should be consistent with the pdf files respectively.
        #     3. worksheet names should be equal to the layout names
        #     4. Bricscad application should be closed.
        #     5. there should be no pdf files with layout names as filenames

        test_filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_dir")
        expected_filepath = os.path.join(Constants.TEST_DIR, "testdata", "test_dir", "testdata1.xlsx")
        self.__common(test_filepath, expected_filepath)
