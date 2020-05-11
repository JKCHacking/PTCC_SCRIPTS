#!/usr/bin/env python
from src.constants import Constants
from src.logger import Logger
from collections import namedtuple
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import datetime
import dateutil.parser


class TimesheetCalculator:
    def __init__(self):
        self.logger = Logger().get_logger()
        self.all_employee_list = []
        self.workbook = Workbook()

    def calculate_time(self, log_date, time_in, time_out):
        time_in = self.convert_12_hours(time_in)
        time_out = self.convert_12_hours(time_out)
        log_date = self.convert_date(log_date)

        noon_time1 = self.convert_12_hours('12:00 PM')
        noon_time2 = self.convert_12_hours('1:00 PM')

        mid_night1 = self.convert_12_hours('12:00 AM')
        mid_night2 = self.convert_12_hours('1:00 AM')

        before_morning1 = self.convert_12_hours('6:00 AM')
        before_morning2 = self.convert_12_hours('7:00 AM')

        after_evening1 = self.convert_12_hours('6:00 PM')
        after_evening2 = self.convert_12_hours('7:00 PM')

        constraint_counter = 0

        if self.is_within_time(time_in, time_out, noon_time1, noon_time2):
            constraint_counter += self.convert_to_hours(noon_time2 - noon_time1)
        if self.is_within_time(time_in, time_out, mid_night1, mid_night2):
            constraint_counter += self.convert_to_hours(mid_night2 - mid_night1)
        if self.is_within_time(time_in, time_out, before_morning1, before_morning2):
            constraint_counter += self.convert_to_hours(before_morning2 - before_morning1)
        if log_date.weekday() >= 5:
            if self.is_within_time(time_in, time_out, after_evening1, after_evening2):
                constraint_counter += self.convert_to_hours(after_evening2 - after_evening1)

        if time_out == self.convert_12_hours('12:00 AM'):
            time_out += datetime.timedelta(days=1)
        total_hours = self.convert_to_hours(time_out - time_in)
        total_hours -= constraint_counter
        return total_hours

    def add_to_list(self, timesheet_entry):
        ret = 1
        found_flag = False
        Work = namedtuple('Work', ['projectName', 'taskName', 'date', 'timeIn', 'timeOut', 'totalHours'])
        Employee = namedtuple('Employee', ['employeeName', 'work'])

        # employee_name,date,project,task,time_in,time_out
        try:
            employee_name = timesheet_entry['employee_name'].strip()
            date = timesheet_entry['date'].strip()
            project = timesheet_entry['project'].strip()
            task = timesheet_entry['task'].strip()
            time_in = timesheet_entry['time_in'].strip()
            time_out = timesheet_entry['time_out'].strip()

            total_hours = self.calculate_time(date, time_in, time_out)
            work_obj = Work(projectName=project, taskName=task, date=date,
                            timeIn=time_in, timeOut=time_out, totalHours=total_hours)

            work_list = [work_obj]
            employee_obj = Employee(employeeName=employee_name, work=work_list)

            if self.all_employee_list:
                for employee in self.all_employee_list:
                    if employee.employeeName == employee_name:
                        employee.work.append(work_obj)
                        found_flag = True
                        break
                if not found_flag:
                    self.all_employee_list.append(employee_obj)
            else:
                self.all_employee_list.append(employee_obj)
        except KeyError:
            ret = -1

        return ret

    def display_all_employess(self):
        print(self.all_employee_list)

    def is_holiday(self, date_str):
        is_holi = False
        date_obj = self.convert_date(date_str)
        date_str = date_obj.strftime('%B %d, %Y')
        if date_str in Constants.HOLIDAY_LIST:
            is_holi = True

        return is_holi

    @staticmethod
    def is_within_time(time_in, time_out, fr_time, to_time):
        is_inside = False
        if time_in < fr_time and time_out > to_time:
            is_inside = True
        elif time_in < fr_time < time_out <= to_time:
            is_inside = True
        elif fr_time <= time_in < to_time < time_out:
            is_inside = True
        elif fr_time < time_in < time_out < to_time:
            is_inside = True

        return is_inside

    @staticmethod
    def convert_12_hours(time_string):
        time_obj = dateutil.parser.parse(time_string)
        time_obj = time_obj.strftime(Constants.TIME_12_FORMAT)
        time_obj = datetime.datetime.strptime(time_obj, Constants.TIME_12_FORMAT)
        return time_obj

    @staticmethod
    def convert_24_hours(time_string):
        time_obj = dateutil.parser.parse(time_string)
        time_obj = time_obj.strftime(Constants.TIME_24_FORMAT)
        time_obj = datetime.datetime.strptime(time_obj, Constants.TIME_24_FORMAT)
        return time_obj

    @staticmethod
    def convert_date(date_string):
        date_obj = None
        try:
            date_obj = dateutil.parser.parse(date_string)
            date_obj = date_obj.strftime(Constants.DATE_FORMAT)
            date_obj = datetime.datetime.strptime(date_obj, Constants.DATE_FORMAT)
        except (TypeError, ValueError):
            print("You have input an Invalid date")

        return date_obj

    @staticmethod
    def convert_to_hours(raw_time):
        (h, m, s) = str(raw_time).split(':')
        time_hours = int(h) + int(m) / 60 + int(s) / 3600
        return time_hours

    @staticmethod
    def insert_table_headers(ws):
        ws.append(['Date', 'Day', 'Time-in Time-out', 'Rendered MINS for the Day',
                   'Credited Regular Log[480 = 1 day]', 'Minutes in excess of 480; Sat/Sun Duties'])

    @staticmethod
    def get_sunday_clusters(fr_day, to_day):
        start = fr_day
        end = to_day
        sunday_cluster_list = []

        while start <= end:
            sunday_list = []
            while start.weekday() != 6:
                sunday_list.append(start)
                start += datetime.timedelta(days=1)
                if start == end:
                    break
            sunday_list.append(start)
            sunday_cluster_list.append(sunday_list)
            start += datetime.timedelta(days=1)

        return sunday_cluster_list

    @staticmethod
    def adjust_cell_height_width(worksheet):
        dims_col = {}
        dims_row = {}

        for i, row in enumerate(worksheet.rows):
            for cell in row:
                if cell.value:
                    dims_col[cell.column_letter] = max((dims_col.get(cell.column_letter, 0), len(str(cell.value))))
                    cell_value = cell.value
                    newline_count = str(cell_value).count('\n')
                    max_value = max((dims_row.get(i+1, 0), newline_count))
                    dims_row[i+1] = max_value*2

        for col, value in dims_col.items():
            worksheet.column_dimensions[col].width = value

        for row, value in dims_row.items():
            if value != 0:
                worksheet.row_dimensions[row].height = value

    def generate_between_days(self, fr_day, to_day):
        fr_day = self.convert_date(fr_day)
        to_day = self.convert_date(to_day)

        if fr_day is None or to_day is None or fr_day > to_day:
            return None

        emp_list = []

        for employee in self.all_employee_list:
            emp_obj = {}
            rendered_hours = 0
            emp_obj['NAME'] = employee.employeeName
            emp_obj['PROJECT'] = []
            for work in employee.work:
                work_date = self.convert_date(work.date)
                if fr_day <= work_date <= to_day:
                    rendered_hours += work.totalHours
                    emp_obj['TIME_SPEND'] = rendered_hours
                    emp_obj['PROJECT'].append(work.projectName)
                else:
                    emp_obj['TIME_SPEND'] = 0
                    emp_obj['PROJECT'].append('None')

            emp_obj['PROJECT'] = set(emp_obj['PROJECT'])
            emp_obj['PROJECT'] = '/'.join(emp_obj['PROJECT'])
            if emp_obj['TIME_SPEND'] != 0:
                emp_list.append(emp_obj)

        date_today = datetime.datetime.today().strftime(Constants.DATE_FORMAT)
        ws = self.workbook.active
        ws.title = "Summary"
        ws["A1"] = f'PTCC TIME SHEET FOR WEEK ({fr_day.date()} to {to_day.date()})'
        ws["A2"] = f'{date_today}'

        data_len = len(emp_list)
        offset = 3
        header = 1
        ws.cell(row=4, column=1, value='NAME')
        ws.cell(row=4, column=2, value='PROJECT')
        ws.cell(row=4, column=3, value='TIME SPEND')

        for row in range(data_len):
            ws.cell(row=row+offset+header+1, column=1, value=emp_list[row]['NAME'])
            ws.cell(row=row+offset+header+1, column=2, value=emp_list[row]['PROJECT'])
            ws.cell(row=row+offset+header+1, column=3, value=emp_list[row]['TIME_SPEND'])

        tabl = Table(displayName='employee_summary', ref=f'A4:C{data_len+offset+header}')
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabl.tableStyleInfo = style
        ws.add_table(tabl)
        ws.append([' '])
        ws.append(['APPROVED by (SUPERVISOR)'])
        ws.append([' '])
        ws.append(['_________________________'])
        ws.append(['ADAM LEE'])
        ws.append([' '])
        ws.append([' '])
        ws.append(['NOTED BY:'])
        ws.append([' '])
        ws.append(['_________________________'])
        ws.append(['SUSAN BENJAMIN'])

        self.adjust_cell_height_width(ws)
        output_path = os.path.join(Constants.OUTPUT_DIR, f'employee_timesheet_{fr_day.date()}_{to_day.date()}.xlsx')
        self.workbook.save(output_path)
        return 1

    def generate_manhour_analysis(self, fr_day, to_day):
        fr_day = self.convert_date(fr_day)
        to_day = self.convert_date(to_day)

        if fr_day is None or to_day is None:
            return None

        for employee in self.all_employee_list:
            ws = self.workbook.create_sheet(f'{employee.employeeName}_summary')
            ws['A1'] = employee.employeeName
            ws['A2'] = 'MANHOUR ANALYSIS OF RENDERED DUTIES'
            ws['A3'] = f'PERIOD COVERED {fr_day.strftime(Constants.DATE_FORMAT)} to ' +\
                       f'{to_day.strftime(Constants.DATE_FORMAT)}'
            ws.append([' '])

            sunday_cluster_list = self.get_sunday_clusters(fr_day, to_day)
            for week in sunday_cluster_list:
                ws.append([f'WEEK COVERED: {week[0].strftime(Constants.DATE_FORMAT)} to ' +
                           f'{week[len(week)-1].strftime(Constants.DATE_FORMAT)}'])
                self.insert_table_headers(ws)
                for day in week:
                    total_hours = 0
                    time_in_out_list = []
                    time_in_out_list_sorted = []
                    for work in employee.work:
                        work_date = self.convert_date(work.date)
                        if day == work_date:
                            time_in_temp = self.convert_24_hours(work.timeIn)
                            time_out_temp = self.convert_24_hours(work.timeOut)
                            time_in_out_list.append(f'{time_in_temp.time()}-{time_out_temp.time()}')
                            total_hours += work.totalHours

                    time_in_out_list.sort()

                    for time_in_out in time_in_out_list:
                        time_split = time_in_out.split('-')
                        time_in_obj = self.convert_12_hours(time_split[0])
                        time_out_obj = self.convert_12_hours(time_split[1])
                        time_in_12_str = time_in_obj.strftime(Constants.TIME_12_FORMAT)
                        time_out_12_str = time_out_obj.strftime(Constants.TIME_12_FORMAT)
                        time_in_out_list_sorted.append(f'{time_in_12_str}-{time_out_12_str}')

                    joined_time_in_out = '\n'.join(time_in_out_list_sorted)
                    day_str = day.strftime(Constants.DATE_FORMAT)
                    total_minutes = total_hours * 60
                    if self.is_holiday(day_str):
                        day_str = f'{day_str}-HOLIDAY'
                        credited_min = 0
                        excess = total_minutes
                    else:
                        credited_min = 480 if total_minutes >= 480 else total_minutes
                        excess = total_minutes - credited_min

                    ws.append([day_str, Constants.DAY_LIST[day.weekday()], joined_time_in_out,
                               total_minutes, credited_min, excess])
                self.adjust_cell_height_width(ws)
                ws.append([' '])

        output_path = os.path.join(Constants.OUTPUT_DIR, f'employee_timesheet_{fr_day.date()}_{to_day.date()}.xlsx')
        self.workbook.save(output_path)
        return 1
