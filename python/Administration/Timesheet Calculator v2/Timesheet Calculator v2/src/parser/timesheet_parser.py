import os
import datetime
import openpyxl
from src.utils.constants import Constants
from src.entity.employee import Employee
from src.entity.task import Task
from src.entity.employee_timesheet import EmployeeTimesheet


class TimesheetParser:
    def iter_spreadsheets(self):
        directory = Constants.INPUT_DIR
        for dirpath, dirnames, filenames in os.walk(directory):
            for filename in filenames:
                fullpath = os.path.join(dirpath, filename)
                if fullpath.endswith(Constants.EXCEL_EXT):
                    yield fullpath

    def parse(self):
        timesheet_list = []
        for excel_fp in self.iter_spreadsheets():
            wb = openpyxl.load_workbook(excel_fp)
            wb_sheetnames = wb.sheetnames
            # check if they are using the correct spreadsheet supported.
            if "secret_identifier" in wb_sheetnames:
                ws = wb[wb_sheetnames[0]]
                id_number = ws["A6"].value
                first_name = ws["B6"].value
                middle_name = ws["C6"].value
                last_name = ws["D6"].value
                # checking of value
                id_number = id_number if id_number is not None else ""
                first_name = first_name if first_name is not None else ""
                middle_name = middle_name if middle_name is not None else ""
                last_name = last_name if last_name is not None else ""

                # create Employee object
                emp = Employee(id_number=id_number,
                               first_name=first_name,
                               middle_name=middle_name,
                               last_name=last_name)
                # create list of Task objects
                task_list = []
                row_num = 11
                while not all(cell.value is None for row in ws["A{}".format(row_num):"E{}".format(row_num)]
                              for cell in row):
                    date = ws["A{}".format(row_num)].value
                    time_in = ws["B{}".format(row_num)].value
                    time_out = ws["C{}".format(row_num)].value
                    task_name = ws["D{}".format(row_num)].value
                    project_name = ws["E{}".format(row_num)].value
                    # checking of value
                    task_name = task_name if task_name is not None else ""
                    project_name = project_name if project_name is not None else ""
                    # date, start_time and end_time must be valid datetime object
                    # to avoid error in the calculator later
                    if isinstance(date, datetime.datetime):
                        date = date.date()
                    if isinstance(time_in, datetime.datetime):
                        time_in = time_in.time()
                    if isinstance(time_out, datetime.datetime):
                        time_out = time_out.time()
                    if isinstance(date, datetime.date) and isinstance(time_in, datetime.time) and \
                            isinstance(time_out, datetime.time):
                        task = Task(date=date,
                                    time_in=time_in,
                                    time_out=time_out,
                                    task_name=task_name,
                                    project_name=project_name)
                        task_list.append(task)
                    row_num += 1
                emp_ts = EmployeeTimesheet(
                    employee=emp,
                    task_list=task_list
                )
                timesheet_list.append(emp_ts)
        return timesheet_list
